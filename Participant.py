from Participanttask import ParticipantTask
from STUDYGROUP import STUDYGROUP
from Task import Task

import SETTINGS as settings


import json
import os
from openpyxl import load_workbook



class Participant(object):

    def __init__(self, id:str, studygroup: STUDYGROUP):
        self.id = id
        self.studygroup = studygroup
        self.tasks = []


    def get_data_directory(self):
        return f"DATA/participant_{self.id}/"

    def get_demographics_filename(self):
        return self.get_data_directory() + f"demographics_{self.id}.json"

    def create_task_files(self, task:Task):

        filename = f"task_{self.id}_{task.task_number}.json"
        ########### USE JSON
        if settings.DATA_INPUT == "JSON":
            with open(self.get_data_directory()+filename, "r", encoding="utf-8") as f:
                raw_data = json.load(f)
                p_task = ParticipantTask(p_id=self.id, task_id=task.task_number)
                p_task.set_raw_data(raw_data)
                self.add_participant_task(p_task=p_task)
            print(f"✓ Loaded from JSON: {filename}")
            return

        ########### READ FROM EXCEL FILE #########
        if settings.DATA_INPUT == "EXCEL":
            output_dir = self.get_data_directory()
            wb = load_workbook("DATA/Data_Study_KEO_Analyzing_Complete.xlsx", data_only=True)
            ws = wb[f"{self.id}({self.studygroup.name})"]

            task_line = task.get_excel_line()
            template = {
                    "Task ID": ws[f"A{task_line}"].value,# just to be sure
                    "Result Cell understood (1 / 0)": {
                        "explicitly mentioned": ws[f"I{task_line}"].value,
                        "implied understand": ws[f"J{task_line}"].value
                    },
                    "each operation understood (1 point for each)": {
                        "explicitly mentioned": ws[f"K{task_line}"].value,
                        "implied understanding": ws[f"L{task_line}"].value
                    },
                    "each expected operand understood (1 point for each)": {
                        "explicitly mentioned": ws[f"M{task_line}"].value,
                        "implied understanding": ws[f"N{task_line}"].value
                    },
                    "Result Type understood (1.0/0.5/0)": {
                        "explicitly mentioned": ws[f"O{task_line}"].value,
                        "implied understanding": ws[f"P{task_line}"].value
                    },
                    "start (of new slide)": str(ws[f"S{task_line}"].value),
                    "timestamp where participant realizes": str(ws[f"T{task_line}"].value),
                    "end of this slide": str(ws[f"U{task_line}"].value),
                    "notes": str(ws[f"V{task_line}"].value)
            }

            os.makedirs(output_dir, exist_ok=True)
            filepath = os.path.join(output_dir, filename)

            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(template, f, indent=2, ensure_ascii=False)

            print(f"! Updated Data from Excel for: {filepath}")
            p_task = ParticipantTask(p_id=self.id, task=task)
            p_task.set_raw_data(template)
            self.add_participant_task(p_task=p_task)
        return

    def add_participant_task(self, p_task:ParticipantTask):
        task_ids = []
        for t in self.tasks:
            task_ids.append(t.id)
        if p_task.id not in task_ids:
            self.tasks.append(p_task)


